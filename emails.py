
import sys
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from openpyxl import load_workbook

teacher_emails = None

user_email = "upper_campus@gmail.com"
app_password = "???" # Use the generated App Password, not your regular password

'''
def email_all_families(email_dict, sample_email):
	#print(email_dict)

	mail_list = [ for i in email_dict]

	for email,children in email_dict.items():
		#print(f'to {email}:')

		cstr = children_list_to_str(children)
		this_email = sample_email.replace('{children}', cstr)
		#print(this_email)
		send_email(email, 'Your Q2 report cards', this_email)
		#for child in children:
		#	print(f'{child}, ')
'''
#send_bulk_emails(data [name, addr, other], lambda, subject, body)
def email_all_teachers(missing_entries):
	global teacher_emails
	teacher_emails = read_email_list('user/teacher_emails.xlsx')
	#print(teacher_emails)


	subject = 'Report cards Qx -- Missing Entries'
	body = 'Dear {teacher},\n\nYou are missing writeups in the following report cards:\n{missing}\nPlease fix these soon!\nSincerely,\nSam'
	def missing_entry_compose_email(body, name, values):
		#global teacher_emails
		#print(name)
		newbody = body.replace("{teacher}", name)
		mstr = ''
		for m in values:
			url = f"https://docs.google.com/document/d/{m[1]}/edit"
			mstr += f'\t{m[0]}\t{url}\n'
		newbody = newbody.replace("{missing}", mstr)
		return newbody

	addresses = []
	for name in missing_entries:
		tchemail = get_teacher_email(name)
		if not tchemail or tchemail.strip() == '':
			print(f'No email found for {name}!')
			sys.exit()
		addresses.append(tchemail)

	send_bulk_emails(missing_entries, missing_entry_compose_email, addresses, subject, body)

def get_teacher_email(name):
	global teacher_emails
	this_email = None
	for item in teacher_emails:
		 if item[0] == name:
				this_email = item[1]
				break
	return this_email

'''
f is a lambda/function that composes each email.
it does this my modifying the "body" param, according to fields in the
"data" item.
'''
#data [name, addr, other], f, subject, body
def send_bulk_emails(data, f, addresses, subject, body):
	for i, (name,values) in enumerate(data.items()):

		#for name,values in data.items():
		body_updated = f(body, name, values)
		#send_email(item.get("addr"), subject, body_updated)
		send_email(addresses[i], subject, body_updated)


#list of structure which contains "addr" and "email"
#def send_bulk_emails(mail_list, subject):
#	for addr,text in mail_list.items():
#		send_email(addr, subject, text)

def send_email(dst_email, subject, body):
	global user_email
	global app_password

	#for now
	print('-' * 40)
	print(f'[From]:\t\t{user_email}')
	print(f'[To:]\t\t{dst_email}')
	print(f'[Subject]:\t{subject}')
	print(f'[Body]:\n{body}')
	print('-' * 40)
	return

	#safeguard until it's well tested
	if not dst_email == samrosenfield22@gmail.com:
		print('only send email to me for now!')
		sys.exit()


	# Create a MIMEText object for the email body
	message = MIMEMultipart()
	message["From"] = user_email
	message["To"] = dst_email
	message["Subject"] = subject
	message.attach(MIMEText(body, "plain"))

	# Connect to Gmail's SMTP server and send the email
	try:
		# Use port 465 for SSL or 587 for STARTTLS
		# Port 465 with SMTP_SSL is often the easiest and most reliable
		server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
		server.login(user_email, app_password)
		server.send_message(message)
		print("Email sent successfully!")
	except Exception as e:
		print(f"Error sending email: {e}")
	finally:
		server.quit() # Close the connection

def children_list_to_str(names):
	cstr = ''
	for child in names:
		if child == names[-1]:
			cstr += child
		elif child == names[len(names)-2]:
			cstr += f'{child} and '
		else:
			cstr += f'{child}, '
	return cstr

#converts the list of lists into a dictionary, with shared
#email users as a list
def email_list_to_dict(path):

	students_and_emails = read_email_list(path)

	mydict = {}
	for v,k in students_and_emails:
		mydict.setdefault(k, [])
		mydict[k].append(v)
	return mydict

def read_email_list(path):
	workbook = load_workbook(filename=path)

	# Select the desired sheet
	sheet = workbook.active

	# Use list comprehension with iter_rows to get all values
	# values_only=True extracts the cell's value directly, not the cell object
	data_as_list = [row for row in sheet.iter_rows(values_only=True)]

	return data_as_list


#d = email_list_to_dict('dummy_emails.xlsx')
#sample_email = 'To whom it may concern:\nReport card(s) for {children} are attached below!\n\nSincerely,\nAdministrator'
#send_bulk_emails(d, sample_email)
