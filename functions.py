#Enable the APIs: In the Google Cloud Console, create a new project and enable the Google Docs API and Google Drive API.
#Create Credentials: Set up an OAuth 2.0 client ID (selecting "Desktop app" is a common quickstart option) and download the credentials.json file to your project directory.
#Install Libraries: Install the required Python packages:

import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
#from googleapicliclient.errors import HttpError
from googleapiclient.errors import HttpError

#output to excel
from openpyxl import Workbook
from openpyxl.styles import Font

#export pdfs
import io
import shutil
from googleapiclient.http import MediaIoBaseDownload

#gui
import customtkinter as ctk

# If modifying these scopes, delete the file token.json.
SCOPES = [
	"https://www.googleapis.com/auth/documents",
	"https://www.googleapis.com/auth/drive"
]
#SCOPES = ["https://www.googleapis.com/auth/drive"]

#ALL_DIRECTORIES = ['1DOHwWe-9nxBvhzzJ5xWsOUoBfL4EGQs9']
ALL_DIRECTORIES = []

drive_service = None
docs_service = None

outdir = 'generated/'
outdir_pdfs = outdir + 'PDFs/'

current_doc_title = ''
current_doc_id = ''

#
missing_entries = {}

font_fixes = 0
expected_font_family = 'Garamond'
expected_font_size = 10

message = None

def authenticate_google_services():
	global drive_service, docs_service

	creds = None
	# The file token.json stores the user's access and refresh tokens, and is
	# created automatically when the authorization flow completes for the first
	# time.
	if os.path.exists("token.json"):
		creds = Credentials.from_authorized_user_file("token.json", SCOPES)
	# If there are no valid credentials available, let the user log in.
	if not creds or not creds.valid:
		if creds and creds.expired and creds.refresh_token:
			creds.refresh(Request())
		else:
			flow = InstalledAppFlow.from_client_secrets_file(
				"credentials.json", SCOPES
			)
			creds = flow.run_local_server(port=0)
		# Save the credentials for the next run
		with open("token.json", "w") as token:
			token.write(creds.to_json())

	try:
		#service = build("docs", "v1", credentials=creds)
		drive_service = build("drive", "v3", credentials=creds)
		docs_service = build("docs", "v1", credentials=creds)

	except HttpError as err:
		print(err)
		#return None

def get_all_folder_names():
	global ALL_DIRECTORIES
	folder_names = []
	for folder_id in ALL_DIRECTORIES:
		if folder_id:
			name = folder_get_name(folder_id)
			folder_names.append(name)
	#print(folder_names)
	return folder_names

def folder_get_name(folder_id):
	global drive_service
	page_token = None
	try:
		file_metadata = drive_service.files().get(
			fileId=folder_id,
			fields='name, mimeType' # Request only the 'name' and 'mimeType' fields to be efficient
		).execute()

		if file_metadata['mimeType'] == 'application/vnd.google-apps.folder':
			return file_metadata['name']
		else:
			return None

	except HttpError as error:
		print(f'An error occurred: {error}')
		return None

	return items

def folder_get_docs(folder_id):
	global drive_service
	items = []
	page_token = None
	try:
		while True:
			# Query to list files where the 'parent' is the specified folder ID
			# and it's not a folder itself (to list 'docs'/files only)
			query = f"'{folder_id}' in parents and mimeType != 'application/vnd.google-apps.folder'"
			results = drive_service.files().list(
				q=query,
				pageSize=10, # Adjust pageSize as needed
				fields="nextPageToken, files(id, name, mimeType)",
				pageToken=page_token,
				supportsAllDrives=True, # Include this if working with shared drives
				includeItemsFromAllDrives=True).execute()

			items.extend(results.get('files', []))
			#print(items)
			page_token = results.get('nextPageToken', None)
			if page_token is None:
				break
	except HttpError as error:
		print(f'An error occurred: {error}')
		return []

	return items

def process_doc(DOCUMENT_ID, op_fix_fonts, op_missing_writeup_report):
	doc = open_doc(DOCUMENT_ID)
	#print(f"The title of the document is: {current_doc_title}")

	metatables = get_tables(doc)
	#print(f"found {len(metatables)} table(s)")

	for t in metatables:
		process_table_elements(t[0],t[1], op_fix_fonts, op_missing_writeup_report)

def missing_entries_report(outdir):
	global missing_entries

	wb = Workbook()
	ws = wb.active
	ws.title = 'missing entries'
	ws.column_dimensions['A'].width = 15
	ws.column_dimensions['B'].width = 35

	print('')
	print('*' * 40)
	print("*	 missing entries report")
	print('*' * 40)
	ws.append(['Teacher', 'Missing reports', 'Link'])
	for teacher,missings in missing_entries.items():
		print(f'{teacher} is missing writeups in {len(missings)} reports:')
		ws.append([teacher])
		for rep_name,rep_id in missings:
			url = f"https://docs.google.com/document/d/{rep_id}/edit"
			print(f'\t{rep_name} ({url})')
			ws.append(['', rep_name, url])

	bold_font = Font(bold=True)
	ws['A1'].font = bold_font
	ws['B1'].font = bold_font
	ws['C1'].font = bold_font

	#save spreadsheet
	outfile = outdir + 'missing entries report.xlsx'
	wb.save(filename=outfile)


def open_doc(DOCUMENT_ID):
	document = docs_service.documents().get(documentId=DOCUMENT_ID).execute()
	global current_doc_id, current_doc_title
	current_doc_id = DOCUMENT_ID
	current_doc_title = document.get('title')
	return document

#def get_text(doc):
#	content = doc.get('body').get('content')
#	text = ""
#	for element in content:
#		if element.get('paragraph'):
#			for el in element.get('paragraph').get('elements'):
#				if el.get('textRun'):
#					text += el.get('textRun').get('content')

def get_tables(document):
	content = document.get('body').get('content')

	all_tables = []
	prev_elem = None
	for element in content:
	#for (i, element) in enumerate(content):
		# Check if the structural element is a table

		if element.get('table') is not None:
			table = element.get('table')
			#hdr_text = prev_elem.get('paragraph').get('elements').get('textRun').get('content')
			hdr_text = prev_elem.get('paragraph').get('elements')[0].get('textRun').get('content')
			teacher = hdr_text.replace('Teacher: ', '').rstrip()
			#print(f"table w hdr: {teacher}")
			#print(table)
			all_tables.append([teacher, table])
		prev_elem = element
	return all_tables

def fix_font(run):
	text_run = run.get('textRun')
	style = text_run.get('textStyle')
	if style:
		#print(style)
		font_family = style.get('weightedFontFamily', {}).get('fontFamily', 'N/A')
		font_size_pt = style.get('fontSize', {}).get('magnitude', 'N/A')
		#font_is_bold = style.get('bold', False)
		#font_is_italic = style.get('italic', False)

		if not((font_family == expected_font_family) and (font_size_pt == expected_font_size)):
			#print(style)
			start = run.get("startIndex")
			end = run.get("endIndex")
			#print(f'run from {start} to {end}')
			send_font_fix_request(start, end)

def send_font_fix_request(start, end):
	global font_fixes

	new_text_style = {
		'weightedFontFamily': {
			'fontFamily': expected_font_family
		},
		#'bold': False,
		#'italic': False,
		'fontSize': {
			'magnitude': expected_font_size,
			'unit': "PT"
		}
	}

	# Define the range of text to format (e.g., from index 1 to 10)
	# Note: Indices are 1-based in the API
	text_range = {
		'startIndex': start,
		'endIndex': end
	}

	# Create the update request
	requests = [
		{
			'updateTextStyle': {
				'range': text_range,
				'textStyle': new_text_style,
				'fields': 'weightedFontFamily,fontSize' # Specify the fields to update
			}
		}
	]

	# Execute the batch update request
	try:
		docs_service.documents().batchUpdate(
			documentId=current_doc_id,
			body={'requests': requests}
		).execute()
		#print(f"Font updated successfully for the specified range in document")
		font_fixes+=1
	except Exception as e:
		print(f"An error occurred: {e}")

def process_table_elements(teacher, table,
	op_fix_fonts, op_missing_writeup_report):

	#global
	global missing_entries
	global current_doc_title

	table_data = []
	current_table_rows = []
	#for row in table.get('tableRows'):
	for r, row in enumerate(table.get('tableRows')):
		current_row_cells = []
		#for cell in row.get('tableCells'):
		for c, cell in enumerate(row.get('tableCells')):
			cell_text = ""
			# A cell can contain multiple content elements (paragraphs, lists, etc.)
			for cell_content in cell.get('content'):
				if 'paragraph' in cell_content:
					for run in cell_content.get('paragraph').get('elements'):
						if 'textRun' in run:
							if op_fix_fonts:
								fix_font(run)

							text_run = run.get('textRun')
							cell_text += text_run.get('content')

			if op_missing_writeup_report:
				#if the entire cell is empty, add to the list of missing teacher entries
				#print(cell_text.strip())
				if r==1 and c==2 and not cell_text.strip():
					#empty cell
					missing_entries.setdefault(teacher, [])
					missing_entries[teacher].append([current_doc_title, current_doc_id])
					#print(f'missing report card text! great job, {teacher}!')

			current_row_cells.append(cell_text.strip())
		current_table_rows.append(current_row_cells)
	table_data.append(current_table_rows)

	return table_data

def export_google_doc_as_pdf(file_id, output_path):
	#print(f"Starting download of file ID: {file_id} as PDF...")

	# Use the export_media method for Google Workspace formats
	request = drive_service.files().export_media(
		fileId=file_id,
		mimeType='application/pdf'
	)

	# Handle the download in chunks
	fh = io.BytesIO()
	downloader = MediaIoBaseDownload(fh, request)
	done = False
	while done is False:
		status, done = downloader.next_chunk()
		#print(f"Download progress: {int(status.progress() * 100)}%")

	# Write the downloaded bytes to the local file
	fh.seek(0)
	with open(output_path, 'wb') as f:
		shutil.copyfileobj(fh, f, length=131072)

	print(f"Successfully downloaded '{output_path}'")

def load_directory_ids():
	with open("directory_ids.txt", "r") as file:
		content = file.read()
	global ALL_DIRECTORIES
	ALL_DIRECTORIES = content.split('\n')
	#print(f'using {len(ALL_DIRECTORIES)} directories')
	#for each_id in dirids:
	#	print(f'dirid: {each_id}')

def process_all_report_cards(op_fix_fonts,
	op_missing_writeup_report):

	reports_ready = False
	print('\n\n\n')

	load_directory_ids()

	#authenticate_google_services()

	for folder_id in ALL_DIRECTORIES:
		#print(f'searching folder: {folder_get_name(folder_id)}')
		docs = folder_get_docs(folder_id)
		#print(f'folder name: {docs[0]['name']}')
		for doc in docs:
			print(f'> reading doc {doc["name"]}')
			process_doc(doc["id"], op_fix_fonts, op_missing_writeup_report)
	print('\nFinished processing all report cards')

	# outputs
	if op_fix_fonts:
		global font_fixes
		print(f'\n\nTotal font fixes: {font_fixes}')


	if op_missing_writeup_report:
		global outdir
		os.makedirs(outdir, exist_ok=True)
		missing_entries_report(outdir)
		if not missing_entries:
			print('\nReports are complete and ready to go out to parents!!!')
			reports_ready = True

	#reports_ready = not missing_entries
	#print(f'missing_entries: {missing_entries}')
	print(f'reports_ready = {reports_ready}')
	return reports_ready

	#if op_generate_pdfs:
	#	make_all_pdfs()

def make_all_pdfs():
	global outdir_pdfs
	os.makedirs(outdir_pdfs, exist_ok=True)
	for folder_id in ALL_DIRECTORIES:
		docs = folder_get_docs(folder_id)
		for doc in docs:
			outfile = outdir_pdfs + doc["name"] + '.pdf'
			export_google_doc_as_pdf(doc["id"], outfile)
	print(f'\nreport PDFs saved to {outdir_pdfs}')
